"""
Core parsing / cleaning / mapping logic for the CRM import converter.
No Flask here - this module is UI-agnostic so it can be tested on its own.
"""
import csv
import io
import re
import unicodedata
from urllib.parse import urlparse, parse_qs, unquote, urlunparse

import openpyxl

# ---------------------------------------------------------------------------
# Target CRM fields shown on the "right side" of the mapping screen.
# ---------------------------------------------------------------------------

TARGET_FIELDS = [
    {'id': 'company_lead', 'label': 'Company Name / Lead Name', 'type': 'text',
     'outputs': ['Company Name', 'Lead Name'],
     'hint': 'Значение попадёт сразу в 2 колонки результата: Company Name и Lead Name'},
    {'id': 'mobile_phone', 'label': 'Mobile Phone', 'type': 'phone', 'outputs': ['Mobile Phone'], 'hint': ''},
    {'id': 'home_phone', 'label': 'Home Phone', 'type': 'phone', 'outputs': ['Home Phone'], 'hint': ''},
    {'id': 'other_phone', 'label': 'Other Phone Number', 'type': 'phone', 'outputs': ['Other Phone Number'], 'hint': ''},
    {'id': 'corporate_website', 'label': 'Corporate Website', 'type': 'url', 'outputs': ['Corporate Website'], 'hint': ''},
    {'id': 'other_website', 'label': 'Other Website', 'type': 'url', 'outputs': ['Other Website'], 'hint': ''},
    {'id': 'work_email', 'label': 'Work E-mail', 'type': 'email', 'outputs': ['Work E-mail'], 'hint': ''},
    {'id': 'home_email', 'label': 'Home E-mail', 'type': 'email', 'outputs': ['Home E-mail'], 'hint': ''},
    {'id': 'other_email', 'label': 'Other E-mail', 'type': 'email', 'outputs': ['Other E-mail'], 'hint': ''},
    {'id': 'country_outreach', 'label': 'Country OUTREACH', 'type': 'text', 'outputs': ['Country OUTREACH'], 'hint': ''},
    {'id': 'outreach_comment', 'label': 'Outreach comment', 'type': 'text', 'outputs': ['Outreach comment'], 'hint': ''},
    {'id': 'comment', 'label': 'Comment', 'type': 'text', 'outputs': ['Comment'], 'hint': ''},
]
TARGET_BY_ID = {t['id']: t for t in TARGET_FIELDS}

OUTPUT_HEADERS = [
    'Company Name', 'Lead Name',
    'Mobile Phone', 'Home Phone', 'Other Phone Number',
    'Corporate Website', 'Other Website',
    'Work E-mail', 'Home E-mail', 'Other E-mail',
    'Country OUTREACH', 'Outreach comment', 'Comment',
    'Source',
]

SOURCE_VALUE = 'import'
MAX_PHONES_PER_CELL = 3

# CSV dialect that matches the known-good ErnestoTRUE.csv reference file:
# ';' delimiter, CRLF line endings, no BOM, quote only when actually needed.
# (The legacy script used ',' + BOM + quote-all, which is why imports looked
# "crooked" - most locales/Bitrix read ';' as the field separator.)
CSV_DELIMITER = ';'
CSV_LINETERMINATOR = '\r\n'
CSV_ENCODING = 'utf-8'


# ---------------------------------------------------------------------------
# generic text cleaning
# ---------------------------------------------------------------------------

def clean_text(val):
    if val is None:
        return ''
    if isinstance(val, float) and val.is_integer():
        val = int(val)
    val = unicodedata.normalize('NFKC', str(val)).strip()
    val = val.replace('\n', ' ').strip()
    val = re.sub(r'\s{2,}', ' ', val)
    return val


# ---------------------------------------------------------------------------
# email
# ---------------------------------------------------------------------------

EMAIL_RE = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')


def extract_emails(val):
    if not val:
        return []
    s = unicodedata.normalize('NFKC', str(val)).strip()
    s = re.sub(r'(@[\w\-]+),([a-zA-Z]{2,4})\b', r'\1.\2', s)
    found = EMAIL_RE.findall(s)
    seen = set()
    out = []
    for e in found:
        key = e.lower()
        if key not in seen:
            seen.add(key)
            out.append(e)
    return out


# ---------------------------------------------------------------------------
# phone
# ---------------------------------------------------------------------------

URL_RE = re.compile(r'https?://\S+')
WAME_RE = re.compile(r'wa\.me/\S+', re.IGNORECASE)
CTRL_ARTIFACT_RE = re.compile(r'(?:_x00[0-9A-Fa-f]{2}_)+')

PHONE_HEADER_KEYWORDS = (
    'phone', 'tel', 'whatsapp', 'wa', 'cel', 'movil', 'móvil',
    'contact', 'number', 'номер', 'телефон', 'фон',
)


def _split_parens(segment):
    extras = []

    def repl(m):
        content = m.group(1)
        if len(re.sub(r'\D', '', content)) >= 7:
            extras.append(content)
            return ' '
        return ' ' + content + ' '

    main = re.sub(r'\(([^)]*)\)', repl, segment)
    return main, extras


def _clean_phone_candidate(c):
    c = c.replace('(', '').replace(')', '')
    c = re.sub(r'^[^\d+]+', '', c)
    c = re.sub(r'[^\d\s+\-]+$', '', c)
    c = re.sub(r'\s{2,}', ' ', c).strip()
    if len(re.sub(r'\D', '', c)) < 5:
        return ''
    return c


def extract_phone_candidates(val):
    """Pull out every phone-like substring from a cell, cleaned but not yet
    '+'-normalized. Order preserved, de-duplicated by digit signature."""
    if not val:
        return []
    s = unicodedata.normalize('NFKC', str(val))
    s = CTRL_ARTIFACT_RE.sub('', s)

    phones = []
    for m in WAME_RE.findall(s):
        num = m.split('/', 1)[1] if '/' in m else ''
        num = re.sub(r'[^\d+]', '', num)
        if num:
            if not num.startswith('+'):
                num = '+' + num
            phones.append(num)

    s = WAME_RE.sub(' ', s)
    s = URL_RE.sub(' ', s)

    for part in re.split(r'[/&,;]', s):
        part = part.strip()
        if not part:
            continue
        main, extras = _split_parens(part)
        for cand in [main] + extras:
            cleaned = _clean_phone_candidate(cand)
            if cleaned:
                phones.append(cleaned)

    seen = set()
    out = []
    for p in phones:
        sig = re.sub(r'\D', '', p)
        if sig and sig not in seen:
            seen.add(sig)
            out.append(p)
    return out


def normalize_phone(raw):
    """Ensure a phone string starts with '+'. Returns (value, status) where
    status is one of 'ok' (already had +), 'fixed' (we added it),
    'invalid' (too few digits to trust)."""
    raw = raw.strip()
    digits = re.sub(r'\D', '', raw)
    if raw.startswith('+'):
        return raw, 'ok'
    if raw.startswith('00'):
        return '+' + raw[2:].strip(), 'fixed'
    if len(digits) >= 8:
        return '+' + raw, 'fixed'
    return raw, 'invalid'


def looks_like_phone_header(header):
    h = header.lower()
    return any(k in h for k in PHONE_HEADER_KEYWORDS)


# ---------------------------------------------------------------------------
# urls (website)
# ---------------------------------------------------------------------------

TRACKING_PARAMS = {
    'utm_source', 'utm_medium', 'utm_content', 'utm_campaign',
    'fbclid', 'igshid', 'app_absent', 'text', 'brid', 'h', 'igsh',
    'utm_term', 'utm_id',
}


def extract_real_url(url):
    if not url:
        return ''
    url = url.strip()
    if url.startswith('mailto:'):
        return ''

    if 'l.facebook.com/l.php' in url:
        try:
            parsed = urlparse(url)
            params = parse_qs(parsed.query)
            if 'u' in params:
                return extract_real_url(params['u'][0])
        except Exception:
            pass

    if 'l.instagram.com' in url:
        try:
            parsed = urlparse(url)
            params = parse_qs(parsed.query)
            if 'u' in params:
                return extract_real_url(unquote(params['u'][0]))
        except Exception:
            pass

    try:
        parsed = urlparse(url)
        if parsed.query:
            params = parse_qs(parsed.query, keep_blank_values=True)
            cleaned = {k: v for k, v in params.items() if k not in TRACKING_PARAMS}
            new_query = '&'.join(f'{k}={v[0]}' for k, v in cleaned.items()) if cleaned else ''
            url = urlunparse(parsed._replace(query=new_query))
    except Exception:
        pass

    return url


def get_url(val, hyp_target=None):
    if hyp_target and hyp_target.startswith('mailto:'):
        hyp_target = None

    val_str = unicodedata.normalize('NFKC', str(val)).strip() if val is not None else ''

    url_to_use = None
    if val_str.startswith('http://') or val_str.startswith('https://'):
        url_to_use = val_str
    elif val_str and '.' in val_str and ' ' not in val_str and '@' not in val_str and len(val_str) > 4:
        url_to_use = hyp_target if hyp_target else ('https://' + val_str)
    elif hyp_target:
        url_to_use = hyp_target

    if not url_to_use:
        return ''

    return extract_real_url(url_to_use)


# ---------------------------------------------------------------------------
# dedupe helpers used when several source columns map to the same target
# ---------------------------------------------------------------------------

def dedupe_join(values):
    seen = set()
    out = []
    for v in values:
        v = (v or '').strip()
        key = v.lower()
        if v and key not in seen:
            seen.add(key)
            out.append(v)
    return ', '.join(out)


def dedupe_join_phones(values):
    kept = []
    for p in values:
        p = (p or '').strip()
        if not p:
            continue
        digits = re.sub(r'\D', '', p)
        if not digits:
            continue
        dup_idx = None
        for idx, (kd, kp) in enumerate(kept):
            if kd == digits:
                dup_idx = idx
                break
            if kd.endswith(digits) and len(kd) - len(digits) <= 3:
                dup_idx = idx
                break
            if digits.endswith(kd) and len(digits) - len(kd) <= 3:
                kept[idx] = [digits, p]
                dup_idx = idx
                break
        if dup_idx is None:
            kept.append([digits, p])
    return ', '.join(p for _, p in kept)


# ---------------------------------------------------------------------------
# reading the uploaded table
# ---------------------------------------------------------------------------

class TableTooLargeError(Exception):
    pass


def read_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    max_col = ws.max_column
    headers = []
    for c in range(1, max_col + 1):
        h = clean_text(ws.cell(1, c).value)
        headers.append(h or f'Column {c}')

    rows = []
    hyperlinks = []
    for r in range(2, ws.max_row + 1):
        raw_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        if all(v is None or str(v).strip() == '' for v in raw_vals):
            continue
        rows.append([clean_text(v) for v in raw_vals])
        hyperlinks.append([
            (ws.cell(r, c).hyperlink.target if ws.cell(r, c).hyperlink else None)
            for c in range(1, max_col + 1)
        ])
    return headers, rows, hyperlinks


def read_csv_bytes(raw_bytes):
    text = None
    for enc in ('utf-8-sig', 'utf-8', 'cp1251', 'latin-1'):
        try:
            text = raw_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw_bytes.decode('utf-8', errors='replace')

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=';,\t')
        delim = dialect.delimiter
    except Exception:
        delim = ';' if sample.count(';') >= sample.count(',') else ','

    reader = csv.reader(io.StringIO(text), delimiter=delim)
    data = [row for row in reader if any(cell.strip() for cell in row)]
    if not data:
        return [], [], []
    headers = [clean_text(h) or f'Column {i + 1}' for i, h in enumerate(data[0])]
    width = len(headers)
    rows = []
    for r in data[1:]:
        r = list(r) + [''] * (width - len(r))
        rows.append([clean_text(v) for v in r[:width]])
    hyperlinks = [[None] * width for _ in rows]
    return headers, rows, hyperlinks


def read_table(path, filename):
    lower = filename.lower()
    if lower.endswith('.xlsx') or lower.endswith('.xlsm'):
        return read_xlsx(path)
    if lower.endswith('.csv') or lower.endswith('.txt'):
        with open(path, 'rb') as f:
            raw = f.read()
        return read_csv_bytes(raw)
    # fall back: try xlsx first, then csv
    try:
        return read_xlsx(path)
    except Exception:
        with open(path, 'rb') as f:
            raw = f.read()
        return read_csv_bytes(raw)


# ---------------------------------------------------------------------------
# phone-column detection + splitting ("scanning" step)
# ---------------------------------------------------------------------------

def detect_phone_columns(headers, rows, sample_size=300):
    n_cols = len(headers)
    sample = rows[:sample_size]
    phone_cols = []
    for c in range(n_cols):
        header_hit = looks_like_phone_header(headers[c])
        nonempty = 0
        hits = 0
        for row in sample:
            cell = row[c] if c < len(row) else ''
            if not cell:
                continue
            nonempty += 1
            if extract_phone_candidates(cell):
                hits += 1
        content_hit = nonempty >= 1 and (hits / nonempty) >= 0.3
        if header_hit or content_hit:
            phone_cols.append(c)
    return phone_cols


def scan_phones(headers, rows, phone_col_indices):
    """For every detected phone column, extract up to MAX_PHONES_PER_CELL
    normalized numbers per row. Returns:
      phone_data: {col_idx: {'per_row': [[num,...], ...], 'max_count': int,
                              'stats': {'ok':n,'fixed':n,'invalid':n}}}
    """
    phone_data = {}
    for c in phone_col_indices:
        per_row = []
        stats = {'ok': 0, 'fixed': 0, 'invalid': 0}
        max_count = 1
        for row in rows:
            cell = row[c] if c < len(row) else ''
            candidates = extract_phone_candidates(cell)[:MAX_PHONES_PER_CELL]
            normalized = []
            for cand in candidates:
                val, status = normalize_phone(cand)
                stats[status] += 1
                normalized.append(val)
            per_row.append(normalized)
            max_count = max(max_count, len(normalized))
        phone_data[c] = {'per_row': per_row, 'max_count': min(max_count, MAX_PHONES_PER_CELL), 'stats': stats}
    return phone_data


def build_source_columns(headers, rows, phone_data):
    """Build the flat list of mappable "source columns" shown in the UI:
    plain columns as-is, phone columns exploded into up to 3 sub-columns."""
    columns = []
    for c, header in enumerate(headers):
        if c in phone_data:
            info = phone_data[c]
            n = info['max_count']
            for sub in range(n):
                key = f'{c}:{sub}'
                label = header if n == 1 else f'{header} — Phone {sub + 1}'
                samples = []
                for pr in info['per_row']:
                    if sub < len(pr) and pr[sub]:
                        samples.append(pr[sub])
                    if len(samples) >= 3:
                        break
                columns.append({
                    'key': key,
                    'label': label,
                    'kind': 'phone',
                    'samples': samples,
                    'stats': info['stats'] if sub == 0 else None,
                })
        else:
            samples = []
            for row in rows:
                v = row[c] if c < len(row) else ''
                if v:
                    samples.append(v)
                if len(samples) >= 3:
                    break
            columns.append({
                'key': str(c),
                'label': header,
                'kind': 'text',
                'samples': samples,
                'stats': None,
            })
    return columns


# ---------------------------------------------------------------------------
# building the final output
# ---------------------------------------------------------------------------

def _value_for_source(key, row_idx, headers, rows, hyperlinks, phone_data):
    """Resolve the raw value for one source key on one row, before target
    -specific processing (email/url/phone extraction)."""
    if ':' in key:
        col_str, sub_str = key.split(':')
        c, sub = int(col_str), int(sub_str)
        per_row = phone_data[c]['per_row']
        vals = per_row[row_idx] if row_idx < len(per_row) else []
        return vals[sub] if sub < len(vals) else '', None
    c = int(key)
    row = rows[row_idx] if row_idx < len(rows) else []
    val = row[c] if c < len(row) else ''
    hyp = None
    if hyperlinks and row_idx < len(hyperlinks):
        hrow = hyperlinks[row_idx]
        hyp = hrow[c] if c < len(hrow) else None
    return val, hyp


def build_output_rows(headers, rows, hyperlinks, phone_data, mapping):
    """mapping: {source_key: target_id}. Returns list of OUTPUT_HEADERS rows."""
    # group source keys by target id
    by_target = {}
    for key, target_id in mapping.items():
        if not target_id:
            continue
        by_target.setdefault(target_id, []).append(key)

    output_rows = []
    for row_idx in range(len(rows)):
        record = {h: '' for h in OUTPUT_HEADERS}
        record['Source'] = SOURCE_VALUE

        for target_id, keys in by_target.items():
            target = TARGET_BY_ID[target_id]
            ttype = target['type']
            collected = []
            for key in keys:
                raw, hyp = _value_for_source(key, row_idx, headers, rows, hyperlinks, phone_data)
                if not raw:
                    continue
                if ':' in key:
                    # already-normalized phone value from the scanning step
                    if ttype == 'phone':
                        collected.append(raw)
                    else:
                        collected.append(clean_text(raw))
                    continue
                if ttype == 'email':
                    collected.extend(extract_emails(raw))
                elif ttype == 'url':
                    u = get_url(raw, hyp)
                    if u:
                        collected.append(u)
                elif ttype == 'phone':
                    for cand in extract_phone_candidates(raw)[:MAX_PHONES_PER_CELL]:
                        val, _status = normalize_phone(cand)
                        collected.append(val)
                else:
                    collected.append(clean_text(raw))

            if not collected:
                continue

            if ttype == 'phone':
                joined = dedupe_join_phones(collected)
            else:
                joined = dedupe_join(collected)

            for out_col in target['outputs']:
                record[out_col] = joined

        output_rows.append([record[h] for h in OUTPUT_HEADERS])

    return output_rows


def write_csv(path, output_rows):
    with open(path, 'w', newline='', encoding=CSV_ENCODING) as f:
        writer = csv.writer(
            f, delimiter=CSV_DELIMITER, lineterminator=CSV_LINETERMINATOR,
            quoting=csv.QUOTE_MINIMAL,
        )
        writer.writerow(OUTPUT_HEADERS)
        writer.writerows(output_rows)
